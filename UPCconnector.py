from http.client import REQUESTED_RANGE_NOT_SATISFIABLE

import requests
import pandas as pd
import re
from urllib.parse import urlparse, urlunparse
import time
import openpyxl
from tqdm import tqdm
# Replace with your actual API key
api_key = "hhs3029invpl11fd7db1vrvvdi7rdn"
endpoint = "https://api.barcodelookup.com/v3/products"
BEST_BUY_PATTERN = re.compile(r'bestbuy\.com', re.I)
HOME_DEPOT_PATTERN = re.compile(r'homedepot\.com', re.I)
WALMART_PATTERN = re.compile(r'walmart\.com', re.I)
AMAZON_PATTERN = re.compile(r'amazon\.com', re.I)
def extract_product_id(link):
    digit_groups = re.findall(r'\d+', link)
    longest_group = max(digit_groups, key=len, default="")
    return longest_group
def format_walmart_url(url):
    parsed_url = urlparse(url)
    modified_path = parsed_url.path.split('&')[0]
    modified_url = urlunparse((parsed_url.scheme, parsed_url.netloc, modified_path, '', '', ''))
    return modified_url
def format_amazon_url(asin):
    return f"https://www.amazon.com/dp/{asin}"
def extract_amazon_asin(url):
    asin_match = re.search(r'/dp/(\w+)/?', url)
    if asin_match:
        return asin_match.group(1)
    return None
def sanitize_link(link):
    # Check if the link contains "&" symbol
    if '&' in link:
        # Remove everything after the "&" symbol
        link = link.split('&')[0]
    return link
last_request_time = 0
club_upcs_detected = []
club_upc_mbi = {}  # Dictionary to store Club UPCs and their MBI Items
def fetch_data_for_upc(upc):
    global last_request_time
    
    current_time = time.time()
    time_since_last_request = current_time - last_request_time
    
    if time_since_last_request < 1:
        sleep_time = 1 - time_since_last_request
        time.sleep(sleep_time)
    
    url = f"{endpoint}?barcode={upc}&formatted=y&key={api_key}"
    response = requests.get(url)
    
    last_request_time = time.time()
    
    return response
def add_club_upcs_to_excel(upc_list, mbi_list):
    try:
        excel_file = "barcode_info.xlsx"
        wb = openpyxl.load_workbook(excel_file)
        ws = wb.active
        last_row = len(ws["A"]) + 1
        for i, upc in enumerate(upc_list):
            mbi_item = mbi_list[i]
            retailer_link = f"costco.com/CatalogSearch?dept=All&keyword={mbi_item}"
            
            ws.cell(row=last_row, column=1, value=upc)
            ws.cell(row=last_row, column=2, value=mbi_item)
            ws.cell(row=last_row, column=3, value="Costco")
            ws.cell(row=last_row, column=4, value=retailer_link)
            ws.cell(row=last_row, column=5, value=upc)
            
            last_row += 1
        
        wb.save(excel_file)
        wb.close()
    except Exception as e:
        print("Error adding Club UPCs to the Excel file:", str(e))
def fetch_urls_for_upcs(upc_list, club_column, mbi_column):
    data = []
    
    with tqdm(total=len(upc_list), unit="UPC") as progress_bar:
        for i, upc in enumerate(upc_list):
            response = fetch_data_for_upc(upc)
            
            if response.status_code == 200:
                product_data = response.json()["products"][0]
                upc = product_data.get("barcode_number", "")
                title = product_data.get("title", "")
                
                if club_column[i] == "Y":
                    club_upcs_detected.append(upc)
                    club_upc_mbi[upc] = mbi_column[i]
                
                retailer_data = product_data.get("stores", [])
                amazon_asin = product_data.get("asin")
                
                if amazon_asin:
                    data.append([upc, title, "Amazon", format_amazon_url(amazon_asin), amazon_asin])
                    print(f"Successfully fetched data for UPC {upc} from Amazon")
                
                for retailer in retailer_data:
                    name = retailer.get("name", "")
                    link = retailer.get("link", "")
                    
                    if name.lower() == "walmart":
                        formatted_link = format_walmart_url(link)
                    else:
                        formatted_link = sanitize_link(link)
                    
                    product_id = extract_product_id(formatted_link)
                    data.append([upc, title, name, formatted_link, product_id])
            else:
                print(f"Failed to fetch data for UPC {upc}")
            
            progress_bar.update(1)
    
    if data:
        df = pd.DataFrame(data, columns=["UPC", "Product Title", "Retailer", "Retailer Link", "Product ID"])
        excel_file = "barcode_info.xlsx"
        df.to_excel(excel_file, index=False)
        print(f"Data saved to {excel_file}")
    else:
        print("No data to save")
    
    if club_upcs_detected:
        print("\nClub UPCs Detected:")
        for upc in club_upcs_detected:
            mbi_item = club_upc_mbi.get(upc, 'N/A')
            print(f"UPC: {upc}, MBI Item: {mbi_item}")
        add_club_upcs_to_excel(club_upcs_detected, [club_upc_mbi.get(upc, 'N/A') for upc in club_upcs_detected])
    
    print("\nProgram Execution Completed!")
def get_rate_limits():
    url = f"https://api.barcodelookup.com/v3/rate-limits?key={api_key}"
    response = requests.get(url)
    
    if response.status_code == 200:
        rate_limits = response.json()
        allowed_calls_per_month = rate_limits.get("allowed_calls_per_month", "N/A")
        remaining_calls_per_month = rate_limits.get("remaining_calls_per_month", "N/A")
        allowed_calls_per_minute = rate_limits.get("allowed_calls_per_minute", "N/A")
        remaining_calls_per_minute = rate_limits.get("remaining_calls_per_minute", "N/A")
        
        print("\nRate Limits:")
        print(f"Allowed Calls Per Month: {allowed_calls_per_month}")
        print(f"Remaining Calls Per Month: {remaining_calls_per_month}")
        print(f"Allowed Calls Per Minute: {allowed_calls_per_minute}")
        print(f"Remaining Calls Per Minute: {remaining_calls_per_minute}")
    else:
        print("\nFailed to fetch rate limits")
def main():
    print("Welcome to the Barcode Lookup Program!")
    
    try:
        df = pd.read_excel("Imported UPCs.xlsx")
        upc_list = df["UPC"].tolist()
        club_column = df["Club"].tolist()
        mbi_column = df["MBI Item"].tolist()
    except Exception as e:
        print("Error loading UPCs, Club column, and MBI Item column from the Excel file.")
        return
    
    if not upc_list:
        print("No UPCs found in the Excel file.")
        return
    
    fetch_urls_for_upcs(upc_list, club_column, mbi_column)
    
    get_rate_limits()
if __name__ == "__main__":
    main()
